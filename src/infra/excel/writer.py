from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl


def _norm_header(h: str) -> str:
    return (h or "").replace("\t", "").strip().upper()


def load_template_headers(ws) -> Dict[str, int]:
    """
    Retorna dict: HEADER_NORMALIZADO -> coluna (1-based)
    """
    mapping: Dict[str, int] = {}
    for col in range(1, 200):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = _norm_header(str(v))
        if key:
            mapping[key] = col
    return mapping


def write_rows_to_template(
    template_path: Path,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    output_path: Path,
) -> None:
    """
    Escreve rows (lista de dicts) no template, a partir da linha 2.
    As chaves do dict devem corresponder aos headers (normalizados).
    """
    wb = openpyxl.load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada no template: {sheet_name}. Disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = load_template_headers(ws)

    start_row = 2
    for i, row in enumerate(rows):
        excel_row = start_row + i
        for k, v in row.items():
            hk = _norm_header(k)
            col = headers.get(hk)
            if not col:
                # se a coluna não existe no template, ignora
                continue
            ws.cell(row=excel_row, column=col).value = v

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        # Arquivo provavelmente aberto no Excel; salva com sufixo _novo
        alt_path = output_path.with_stem(output_path.stem + "_novo")
        wb.save(alt_path)
        raise PermissionError(
            f"Não foi possível salvar em '{output_path}' (arquivo aberto?). "
            f"Salvo em '{alt_path}' como alternativa."
        )